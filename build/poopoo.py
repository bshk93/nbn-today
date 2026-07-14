#!/usr/bin/env python3
"""
Diffs nbn.today's live cap/roster data against the league's published
Google Sheet (the historical roster/cap-tracking spreadsheet, which the
committee is still hand-updating in parallel this season).

Sheet is fetched via its "Publish to the web" xlsx export, which
auto-updates a few minutes after any edit -- not a frozen snapshot.

Compares, per team, for the site's current league year (from
/api/league-year, expected to match the sheet's "current" column):
  - Aggregate cap figures (guaranteed salary, hard-cap flag) -- sheet
    rows 39-49, sourced the same way cap-summary/ computes them.
  - MLE type/amount/remaining -- sheet G75:J77.
  - BAE year-used -- sheet's "Bi-Annual Exception -> Year Used" cell.
  - TPE remaining -- sheet's "Traded Player Exception -> Exception
    Amount Remaining" cell vs /api/trade-exceptions.
  - Draft picks -- sheet's "Original Draft Picks" table (year/round/
    current owner) vs the global /api/picks ledger.
  - Player-level: every named row in the sheet's "Players" section is
    matched by name (same team first, then fuzzy, then league-wide) to
    a site roster entry, and the two are compared for real-salary
    amount, cap-hold amount, and "hold vs signed" status. Cell fill
    color is read to tell a cap hold apart from a real contract (see
    LEGEND_COLORS) -- reading raw values without this produces false
    conclusions, since the sheet lists holds in the same table as real
    contracts and only distinguishes them by row color.

Writes a single JSON snapshot to $NBS_DATA_DIR/poopoo.json for the
/poopoo/ page to render. Run on a timer (see poopoo.timer) aligned to
the clock (:00/:10/:20/...) so refreshes are predictable.

Env:
  NBN_API_BASE   override the API base URL (default http://127.0.0.1:8001)
  NBS_DATA_DIR   override the data dir (default /var/lib/nothing-but-stats)
  POOPOO_SHEET_URL  override the published-sheet xlsx URL
"""
import csv
import difflib
import io
import json
import os
import re
import unicodedata
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import requests

REPO = Path(__file__).resolve().parent.parent
DATA_DIR = Path(os.environ.get("NBS_DATA_DIR", "/var/lib/nothing-but-stats"))
API_BASE = os.environ.get("NBN_API_BASE", "http://127.0.0.1:8001")
OUT_FILE = DATA_DIR / "poopoo.json"
SHEET_URL = os.environ.get(
    "POOPOO_SHEET_URL",
    "https://docs.google.com/spreadsheets/d/e/2PACX-1vTqyqqC0O1U9O-2uwmMk6UIhSO58ukTa5HpXaU_IOQa3SEW8bLK5Wpjh_KA4YWePDgT2BIdhPO6Mieu/pub?output=xlsx",
)

TEAMS = [
    "ATL", "BKN", "BOS", "CHA", "CHI", "CLE", "DAL", "DEN", "DET", "GSW",
    "HOU", "IND", "LAC", "LAL", "MEM", "MIA", "MIL", "MIN", "NOP", "NYK",
    "OKC", "ORL", "PHI", "PHX", "POR", "SAC", "SAS", "TOR", "UTA", "WAS",
]

DOLLAR_TOLERANCE = 1000  # ignore diffs smaller than this (rounding noise)
FUZZY_NAME_THRESHOLD = 0.84  # difflib ratio for typo-tolerant same-team matching

# Known naming mismatches between the two sources that no generic normalizer
# fixes -- either a nickname (site uses the legal/roster name, sheet uses a
# common nickname) or a family-name-first convention the "last token is the
# surname" heuristic gets backwards. Keyed by sheet display name -> the
# normalized key it should resolve to (see normalize_name).
NAME_ALIASES = {
    "BUB CARRINGTON": "CARRINGTON CARLTON",
    "YANG HANSEN": "YANG HANSEN",  # family name first; skip the last-token-as-surname flip
    "CAMERON THOMAS": "THOMAS CAM",
    "BONES HYLAND": "HYLAND NAHSHON",
    "OLIVIER-MAXENCE PROSPER": "PROSPER OMAX",
    "YANIC KONAN NIEDERHAUSER": "KONAN NIEDERHAUSER YANIC",
    "MO GUEYE": "GUEYE MOUHAMED",
}

# Lowercase name particles that stay glued to the following word as part of a
# compound surname (e.g. "Tristan da Silva" -> surname "da Silva", not "Silva"
# alone) -- checked against the original-case token before upper-casing.
SURNAME_PARTICLES = {"da", "de", "van", "der", "von", "la", "le", "di", "dos", "del"}


def http_json(path):
    with urllib.request.urlopen(f"{API_BASE}{path}", timeout=30) as r:
        return json.loads(r.read().decode())


def parse_csv(text):
    return list(csv.DictReader(io.StringIO(text)))


def parse_salary(s):
    if not s:
        return 0
    try:
        return int(str(s).replace("$", "").replace(",", "").strip())
    except ValueError:
        return 0


# ── Name normalization ──────────────────────────────────────────────────

SUFFIXES = {"JR", "SR", "II", "III", "IV", "V"}


def normalize_name(name):
    """Maps both 'LAST, FIRST' (site) and 'First Last[ Jr.]' (sheet) to a
    common 'LAST FIRST' key, stripping accents/punctuation/suffixes so the
    two sources can be matched by name."""
    if not name:
        return ""
    alias = NAME_ALIASES.get(name.strip().upper())
    if alias:
        return alias
    name = "".join(c for c in unicodedata.normalize("NFKD", name) if not unicodedata.combining(c))
    clean = name.upper().replace(".", "").replace("'", "").replace("’", "")
    if "," in clean:
        last, first = clean.split(",", 1)
        return f"{last.strip()} {first.strip()}"
    raw_toks = name.replace(".", "").replace("'", "").replace("’", "").split()
    toks = clean.split()
    while toks and toks[-1] in SUFFIXES:
        toks.pop()
        raw_toks.pop()
    if len(toks) < 2:
        return " ".join(toks)
    surname_start = len(toks) - 1
    while surname_start > 0 and raw_toks[surname_start - 1].lower() in SURNAME_PARTICLES:
        surname_start -= 1
    surname = " ".join(toks[surname_start:])
    first = " ".join(toks[:surname_start])
    return f"{surname} {first}"


# ── Sheet-side extraction ───────────────────────────────────────────────

def classify(bio, season):
    """Mirrors cap-summary/index.html's classify() -- must stay in sync."""
    g = (bio.get("guaranteed") or {}).get(season)
    if g in ("0", 0):
        return "nongtd"
    hold = (bio.get("cap_holds") or {}).get(season)
    if hold == "PLAYER_OPT":
        return "player_opt"
    if hold == "TEAM_OPT":
        return "team_opt"
    if hold == "RFA":
        return "rfa"
    if hold == "UFA":
        return "ufa"
    return "guaranteed"


def sheet_aggregate(ws, col=6):
    """Rows 39-49, column F (=season col 1, index 6) by default."""
    def cell(r):
        v = ws.cell(row=r, column=col).value
        return v if isinstance(v, (int, float)) else 0

    hard_cap_text = ws.cell(row=47, column=col).value
    hard_cap = None
    if hard_cap_text:
        t = str(hard_cap_text).upper()
        if "SECOND" in t:
            hard_cap = "second_apron"
        elif "FIRST" in t:
            hard_cap = "first_apron"

    return {
        "guaranteed_salary": cell(39),
        "cap_holds": cell(40),
        "cap_space": cell(41),
        "salary_cap": cell(42),
        "first_apron_space": cell(43),
        "first_apron": cell(44),
        "second_apron_space": cell(45),
        "second_apron": cell(46),
        "hard_cap": hard_cap,
        "nbn_hard_cap_space": cell(48),
        "nbn_hard_cap": cell(49),
    }


def sheet_mle(ws):
    return {
        "type": ws.cell(row=76, column=7).value,        # G76
        "amount": ws.cell(row=76, column=9).value or 0,  # I76
        "remaining": ws.cell(row=77, column=9).value or 0,  # I77
    }


def sheet_bae(ws):
    for r in range(78, 96):
        if ws.cell(row=r, column=7).value == "Year Used":
            return ws.cell(row=r, column=9).value
    return None


def sheet_tpe(ws):
    """Traded Player Exception amount remaining -- single slot, one row below
    the 'Exception Amount Remaining' label (col G/I), found dynamically since
    it shifts a row depending on how many draft-pick rows precede it."""
    for r in range(90, 100):
        if ws.cell(row=r, column=7).value == "Exception Amount Remaining":
            return ws.cell(row=r, column=9).value or 0
    return 0


def sheet_picks(ws, orig_team):
    """Original Draft Picks table: col A=Year, B=Round, D=Current Owner.
    Year is merged/blank on the 2nd-round row beneath each 1st-round row.
    Stops at the "Acquired Draft Picks" table, which reuses the same columns
    with different semantics (col D = the *origin* team of a pick BKN etc.
    acquired) -- not comparable to this table and must not be scanned into."""
    picks = []
    last_year = None
    for r in range(76, 160):
        label = ws.cell(row=r, column=1).value
        if isinstance(label, str) and "acquired" in label.lower():
            break
        year_cell = ws.cell(row=r, column=1).value
        round_cell = ws.cell(row=r, column=2).value
        owner_cell = ws.cell(row=r, column=4).value
        if round_cell not in ("1st", "2nd"):
            continue
        if year_cell:
            last_year = year_cell
        if last_year is None or not owner_cell:
            continue
        rnd = 1 if round_cell == "1st" else 2
        owner_raw = str(owner_cell).strip().rstrip("*").strip()
        if "/" in owner_raw:
            continue  # compound/conditional destination (e.g. "CHA/CHI") -- not diffable against a single site owner
        owner = orig_team if owner_raw.lower() == "own" else owner_raw.upper()
        picks.append({"year": int(float(last_year)), "round": rnd, "owner": owner})
    return picks


def build_legend(ws):
    """Reads the Legend row (row 3, cols 6-9) to map fill color -> hold type,
    rather than hardcoding RGB values -- self-adjusts if the sheet's palette
    ever changes. Only UFA/RFA are treated as true cap holds for comparison
    purposes; Player Option/Team Option cells carry real conditional salary
    (still counted in Guaranteed Salary on both sides) so they're compared
    as ordinary contracts, not holds."""
    legend = {}
    for c in range(6, 10):
        label = ws.cell(row=3, column=c).value
        fill = ws.cell(row=3, column=c).fill
        if not label or fill.fill_type != "solid":
            continue
        rgb = fill.fgColor.rgb if fill.fgColor else None
        if not rgb:
            continue
        label_u = str(label).upper()
        if "UFA" in label_u:
            legend[rgb] = "UFA"
        elif "RFA" in label_u:
            legend[rgb] = "RFA"
    return legend


def sheet_players(ws, legend):
    """Parses the main 'Players' section (row 5 through 'Two-Way Contracts').
    Returns a list of {name, salary, hold, pick_num} dicts. `hold` is 'UFA',
    'RFA', or None (real/conditional-real contract, see build_legend).
    'Pick#N' placeholder rows (an unsigned pick's rookie-scale cap hold, no
    real name yet) and 'Empty Roster Charge' rows (a roster-spot hold, not a
    player at all) are surfaced separately from named players -- pick_num is
    set for the former, both are skipped for name-matching purposes."""
    rows = []
    for r in range(5, 37):
        label = ws.cell(row=r, column=1).value
        if isinstance(label, str) and ("two-way" in label.lower() or "dead cap" in label.lower()):
            break
        if not label:
            continue
        name = str(label).strip()
        cell = ws.cell(row=r, column=6)
        salary = cell.value if isinstance(cell.value, (int, float)) else 0
        fill = cell.fill
        rgb = fill.fgColor.rgb if fill.fill_type == "solid" and fill.fgColor else None
        hold = legend.get(rgb)
        pick_match = re.match(r"pick\s*#\s*(\d+)", name, re.IGNORECASE)
        if pick_match:
            rows.append({"name": name, "salary": salary, "hold": hold, "pick_num": int(pick_match.group(1))})
        elif name.lower() == "empty roster charge":
            continue
        else:
            rows.append({"name": name, "salary": salary, "hold": hold, "pick_num": None})
    return rows


def load_sheet():
    import tempfile
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", prefix="poopoo-sheet-")
    os.close(fd)
    tmp = Path(tmp_path)
    resp = requests.get(SHEET_URL, timeout=60)
    resp.raise_for_status()
    tmp.write_bytes(resp.content)
    # read_only=False: need real (non-cached-value-only) cell formatting to
    # read fill colors for cap-hold detection -- costs ~15-20s over read_only
    # for the full 30-team workbook, fine for a 10-minute timer.
    wb = openpyxl.load_workbook(tmp, read_only=False, data_only=True)
    result = {}
    legend = None
    for team in TEAMS:
        if team not in wb.sheetnames:
            continue
        ws = wb[team]
        if legend is None:
            legend = build_legend(ws)
        result[team] = {
            "aggregate": sheet_aggregate(ws),
            "mle": sheet_mle(ws),
            "bae_year_used": sheet_bae(ws),
            "tpe_remaining": sheet_tpe(ws),
            "picks": sheet_picks(ws, team),
            "players": sheet_players(ws, legend or {}),
        }
    tmp.unlink(missing_ok=True)
    return result


# ── Site-side extraction ────────────────────────────────────────────────

def load_site(season, players):
    cap_levels = http_json("/api/cap-levels")
    team_state = http_json("/api/team-state")
    all_picks = http_json("/api/picks")
    trade_exceptions = http_json("/api/trade-exceptions")
    levels = cap_levels.get(season, {})

    result = {}
    for team in TEAMS:
        try:
            roster_csv = (REPO / "data" / f"{team.lower()}-roster.csv").read_text()
            roster_rows = [r for r in parse_csv(roster_csv) if r.get("SLUG")]
        except Exception:
            roster_rows = []
        try:
            dead_rows = http_json(f"/api/deadcap/{team}")
        except Exception:
            dead_rows = []

        cats = {k: 0 for k in ("guaranteed", "dead", "nongtd", "team_opt", "player_opt", "rfa", "ufa")}
        team_players = []
        for r in roster_rows:
            bio = players.get(r["SLUG"])
            if not bio:
                continue
            is_dead = bio.get("type") == "dead" or r.get("TYPE") == "dead"
            salary = parse_salary((bio.get("salaries") or {}).get(season))
            if is_dead or bio.get("type") == "two-way":
                # Two-way players live in the sheet's separate "Two-Way Contracts"
                # table (not parsed for comparison here -- salaries are $0/nominal
                # on both sides, not cap-relevant) so they'd only ever show up as
                # false player_extra noise against the main Players section.
                continue
            if bio.get("type") == "draft-rights":
                team_players.append({
                    "slug": r["SLUG"], "name": bio.get("name", ""), "salary": 0,
                    "hold": None, "cat": "draft-rights",
                    "draft_pick": bio.get("draft_pick"), "draft_year": bio.get("draft_year"),
                })
                continue
            if not salary:
                continue
            cat = classify(bio, season)
            cats[cat] += salary
            team_players.append({
                "slug": r["SLUG"], "name": bio.get("name", ""), "salary": salary,
                "hold": "UFA" if cat == "ufa" else ("RFA" if cat == "rfa" else None),
                "cat": cat, "draft_pick": bio.get("draft_pick"), "draft_year": bio.get("draft_year"),
            })
        for r in dead_rows:
            cats["dead"] += parse_salary(r.get(season))

        guaranteed_salary = cats["guaranteed"] + cats["dead"] + cats["nongtd"] + cats["team_opt"] + cats["player_opt"]
        cap_holds = cats["rfa"] + cats["ufa"]
        total = guaranteed_salary + cap_holds

        ts_cur = (team_state.get(team) or {}).get("current", {})

        result[team] = {
            "aggregate": {
                "guaranteed_salary": guaranteed_salary,
                "cap_holds": cap_holds,
                "cap_space": (levels.get("cap") or 0) - total,
                "salary_cap": levels.get("cap") or 0,
                "first_apron_space": (levels.get("apron1") or 0) - total,
                "first_apron": levels.get("apron1") or 0,
                "second_apron_space": (levels.get("apron2") or 0) - total,
                "second_apron": levels.get("apron2") or 0,
                "hard_cap": ts_cur.get("hard_cap"),
                "nbn_hard_cap_space": (levels.get("hard_cap") or 0) - total,
                "nbn_hard_cap": levels.get("hard_cap") or 0,
            },
            "mle": {
                "type": ts_cur.get("mle_type"),
                "used": ts_cur.get("mle_used", 0),
            },
            "bae_available": (team_state.get(team) or {}).get("bae_available"),
            "tpe_remaining": sum(e.get("remaining", 0) for e in trade_exceptions.get(team, [])),
            "players": team_players,
        }

    picks_by_orig = {}
    for p in all_picks:
        picks_by_orig.setdefault(p["orig"], []).append(p)
    return result, picks_by_orig


# ── Diffing ──────────────────────────────────────────────────────────────

AGGREGATE_FIELDS = [
    ("guaranteed_salary", "Guaranteed Salary"),
    ("hard_cap", "Hard Cap"),
]
# Cap Holds / Cap Space / First & Second Apron Space are deliberately excluded:
# cap-summary's own site-side total doesn't yet track roster-spot holds or
# unsigned-draft-pick holds (see its CATEGORIES comment), so those fields
# mismatch for ~29/30 teams every single run regardless of anything actually
# being wrong -- pure noise until that's tracked as real site-side data.
# The player-level checks below (name-matched) sidestep this gap entirely,
# since roster-spot/unsigned-pick holds never appear as named rows anyway.


def money_diff(a, b):
    return abs((a or 0) - (b or 0)) > DOLLAR_TOLERANCE


def fmt_money(n):
    return f"${n:,.0f}"


def describe_player(p):
    """'UFA $2,450,000' / 'RFA $19,269,273' / 'signed $16,200,000' -- used
    wherever only one side has a name-matched entry, so a player_missing/
    player_extra diff shows what status the side that DOES have them gave
    them, not just a bare dollar figure."""
    if p["hold"]:
        return f"{p['hold']} {fmt_money(p['salary'])}"
    return f"signed {fmt_money(p['salary'])}"


def build_site_name_index(site_data):
    """slug/team lookup by normalized name, across the whole league, for
    cross-team conflict detection (a player the sheet and site disagree
    about which team holds)."""
    index = {}
    for team, data in site_data.items():
        for p in data["players"]:
            if p["cat"] == "draft-rights":
                continue
            index.setdefault(normalize_name(p["name"]), []).append((team, p))
    return index


def build_sheet_name_index(sheet_data):
    index = {}
    for team, data in sheet_data.items():
        for p in data["players"]:
            if p["pick_num"] is not None:
                continue
            index.setdefault(normalize_name(p["name"]), []).append((team, p))
    return index


def find_fuzzy(name_key, candidates_by_key):
    """Best fuzzy match among a dict's keys, for typo-tolerant matching
    (e.g. sheet 'Read' vs site 'Reed'). Returns the matched key or None."""
    best, best_ratio = None, FUZZY_NAME_THRESHOLD
    for key in candidates_by_key:
        ratio = difflib.SequenceMatcher(None, name_key, key).ratio()
        if ratio > best_ratio:
            best, best_ratio = key, ratio
    return best


def diff_players(team, sheet_players_list, site_players_list, site_name_index, sheet_name_index):
    diffs = []
    site_by_key = {}
    for p in site_players_list:
        if p["cat"] == "draft-rights":
            continue
        site_by_key.setdefault(normalize_name(p["name"]), []).append(p)

    matched_site_keys = set()

    for sp in sheet_players_list:
        if sp["pick_num"] is not None:
            continue  # handled by diff_picks_signed
        key = normalize_name(sp["name"])
        site_matches = site_by_key.get(key)
        matched_key = key
        if not site_matches:
            fuzzy_key = find_fuzzy(key, site_by_key)
            if fuzzy_key:
                site_matches = site_by_key.get(fuzzy_key)
                matched_key = fuzzy_key

        if site_matches:
            matched_site_keys.add(matched_key)
            wp = site_matches[0]
            sheet_is_hold = sp["hold"] is not None
            site_is_hold = wp["hold"] is not None
            if sheet_is_hold != site_is_hold:
                diffs.append({
                    "category": "player_status", "field": sp["name"],
                    "sheet": f"{sp['hold']} hold {fmt_money(sp['salary'])}" if sheet_is_hold else f"signed {fmt_money(sp['salary'])}",
                    "site": f"{wp['hold']} hold {fmt_money(wp['salary'])}" if site_is_hold else f"signed {fmt_money(wp['salary'])}",
                })
            elif sheet_is_hold and site_is_hold:
                if wp["salary"] <= 1 and sp["salary"] <= 1:
                    pass  # neither side has a real figure -- nothing to flag either way
                elif wp["salary"] <= 1:
                    # Site has never computed a real hold value for this player at
                    # all (a literal "$1" placeholder, not a competing formula) --
                    # a systemic gap, not a per-player dispute. Distinct category
                    # so it doesn't read as dozens of individual financial disagreements.
                    diffs.append({
                        "category": "player_hold_uncalculated", "field": sp["name"],
                        "sheet": f"{sp['hold']} {fmt_money(sp['salary'])}",
                        "site": "not yet calculated",
                    })
                elif sp["hold"] != wp["hold"] or money_diff(sp["salary"], wp["salary"]):
                    diffs.append({
                        "category": "player_hold", "field": sp["name"],
                        "sheet": f"{sp['hold']} {fmt_money(sp['salary'])}",
                        "site": f"{wp['hold']} {fmt_money(wp['salary'])}",
                    })
            else:
                if money_diff(sp["salary"], wp["salary"]):
                    diffs.append({
                        "category": "player_salary", "field": sp["name"],
                        "sheet": fmt_money(sp["salary"]), "site": fmt_money(wp["salary"]),
                    })
            continue

        # Not on this team's site roster at all -- check league-wide.
        elsewhere = site_name_index.get(key) or site_name_index.get(find_fuzzy(key, site_name_index) or "")
        elsewhere = [(t, p) for t, p in (elsewhere or []) if t != team]
        if elsewhere:
            other_team, _ = elsewhere[0]
            diffs.append({
                "category": "player_team_conflict", "field": sp["name"],
                "sheet": team, "site": other_team,
            })
        else:
            diffs.append({
                "category": "player_missing", "field": sp["name"],
                "sheet": describe_player(sp), "site": "not found",
            })

    # Site players never matched above: either league-wide-sheet-elsewhere
    # (already reported from the other team's perspective, skip) or truly
    # absent from the sheet everywhere (report as extra).
    for wp in site_players_list:
        if wp["cat"] == "draft-rights":
            continue
        key = normalize_name(wp["name"])
        if key in matched_site_keys:
            continue
        in_sheet_anywhere = sheet_name_index.get(key) or sheet_name_index.get(find_fuzzy(key, sheet_name_index) or "")
        if in_sheet_anywhere:
            continue  # cross-team conflict already surfaced from the sheet-side pass
        if wp["salary"] <= 1:
            continue  # a placeholder hold with no real dollar value isn't worth flagging as "extra"
        diffs.append({
            "category": "player_extra", "field": wp["name"],
            "sheet": "not found", "site": f"{team} {describe_player(wp)}",
        })

    return diffs


def diff_picks_signed(team, sheet_players_list, site_players_list, current_draft_year):
    """Catches a recurrence of the 2026 FRP auto-sign mistake: a sheet
    'Pick#N' placeholder (still unsigned) whose site-side counterpart has
    already been fully signed to a contract. Filtered to the current draft
    class only -- draft_pick numbers repeat across years (e.g. every
    draft has a #1 overall), so matching on pick number alone conflates
    this year's still-unsigned pick with an unrelated past #1 overall
    already on the roster (found via WAS's Pick#1 colliding with Alex
    Sarr, the real 2024 #1 overall, before this filter was added)."""
    diffs = []
    by_pick = {
        p["draft_pick"]: p for p in site_players_list
        if p.get("draft_pick") is not None and p.get("draft_year") == current_draft_year
    }
    for sp in sheet_players_list:
        if sp["pick_num"] is None:
            continue
        wp = by_pick.get(sp["pick_num"])
        if wp and wp["cat"] not in ("draft-rights",):
            diffs.append({
                "category": "pick_signed", "field": f"Pick#{sp['pick_num']}",
                "sheet": "unsigned (draft rights)",
                "site": f"{wp['name']} signed {fmt_money(wp['salary'])}",
            })
    return diffs


def diff_team(team, sheet, site, picks_by_orig, site_name_index, sheet_name_index, current_draft_year):
    diffs = []

    s_agg, w_agg = sheet["aggregate"], site["aggregate"]
    for key, label in AGGREGATE_FIELDS:
        sv, wv = s_agg.get(key), w_agg.get(key)
        if key == "hard_cap":
            if sv != wv:
                diffs.append({"category": "aggregate", "field": label, "sheet": sv, "site": wv})
        elif money_diff(sv, wv):
            diffs.append({"category": "aggregate", "field": label, "sheet": sv, "site": wv})

    # MLE: only flag if usage differs meaningfully (types are recomputed live
    # site-side once used=0, so an unused/unused pair is never a mismatch).
    s_mle_used = (s_mle := sheet["mle"])["amount"] - s_mle["remaining"]
    w_mle_used = site["mle"]["used"] or 0
    if money_diff(s_mle_used, w_mle_used):
        diffs.append({"category": "mle", "field": "MLE Used", "sheet": s_mle_used, "site": w_mle_used})

    # BAE: sheet's Year Used implies unavailable the following year.
    s_bae_recent = bool(sheet["bae_year_used"])
    w_bae_avail = site["bae_available"]
    if s_bae_recent and w_bae_avail:
        diffs.append({"category": "bae", "field": "BAE Availability", "sheet": f"used {sheet['bae_year_used']}", "site": "available"})

    # TPE: sheet only has room for a single exception slot per team, so this
    # compares the sheet's one value against the sum of the site's (possibly
    # multiple) trade-exception entries for that team.
    if money_diff(sheet["tpe_remaining"], site["tpe_remaining"]):
        diffs.append({
            "category": "tpe", "field": "TPE Remaining",
            "sheet": sheet["tpe_remaining"], "site": site["tpe_remaining"],
        })

    # Draft picks: compare (year, round) -> owner for this team's own-origin picks.
    site_picks = {(p["year"], p["round"]): p["owner"] for p in picks_by_orig.get(team, [])}
    for p in sheet["picks"]:
        key = (p["year"], p["round"])
        site_owner = site_picks.get(key)
        if site_owner is None:
            continue  # sheet has a future pick beyond the site's tracked horizon
        if site_owner != p["owner"]:
            diffs.append({
                "category": "picks", "field": f"{key[0]} R{key[1]}",
                "sheet": p["owner"], "site": site_owner,
            })

    diffs.extend(diff_players(team, sheet["players"], site["players"], site_name_index, sheet_name_index))
    diffs.extend(diff_picks_signed(team, sheet["players"], site["players"], current_draft_year))

    return diffs


def main():
    season = http_json("/api/league-year")["current_season"]
    players = http_json("/api/players")
    sheet_data = load_sheet()
    site_data, picks_by_orig = load_site(season, players)

    site_name_index = build_site_name_index(site_data)
    sheet_name_index = build_sheet_name_index(sheet_data)
    current_draft_year = 2000 + int(season.split("-")[0])

    teams_out = []
    for team in TEAMS:
        if team not in sheet_data or team not in site_data:
            continue
        diffs = diff_team(team, sheet_data[team], site_data[team], picks_by_orig, site_name_index, sheet_name_index, current_draft_year)
        teams_out.append({
            "team": team,
            "diff_count": len(diffs),
            "diffs": diffs,
            "sheet": sheet_data[team],
            "site": site_data[team],
        })

    out = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "season": season,
        "teams": teams_out,
    }
    OUT_FILE.write_text(json.dumps(out, indent=2))
    print(f"poopoo: wrote {OUT_FILE} ({sum(t['diff_count'] for t in teams_out)} diffs across {len(teams_out)} teams)")


if __name__ == "__main__":
    main()
