#!/usr/bin/env python3
"""Populate `rookie-scale.json` in NBS_DATA_DIR from the league sheet.

§ 7.1 prices a first-round pick's contract off a per-slot scale that indexes to
the cap each year. The site has always had a home for that table
(`GET/PUT /api/rookie-scale`, the `/rookie-scale` page) and it had never been
filled, so every pick signing was typed in by hand off the spreadsheet.

**Read columns F-J, never the NBA base columns beside them.** Each
"{year} Rookie Contracts" tab carries the live NBN figures in F-J and the
underlying real-NBA scale in L-P, and the L-P block is stale on the 2026 tab —
it still holds 2025's numbers, so re-deriving 120% x base there produces last
year's contract for every pick. F-J is the only trustworthy source.

**Five figures per pick, not four.** Y1 and Y2 are guaranteed, Y3 and Y4 are
team options (§ 7.1), and Y5 is not a contract year at all — it's the § 3.10
RFA cap hold the deal rolls into, at 250% of Y4 for picks 1-2 and 300% for
picks 3-30. Storing it here keeps the whole rookie deal, hold included, in one
place; `cap_holds` marks it RFA so nothing mistakes it for salary.

Season keys are derived from the draft year (2026 -> 26-27 .. 30-31) rather
than read from the sheet's header row. The headers were stale by a year once
already, and the derivation is checked against real signed contracts below.

Usage:
    build/venv/bin/python3 build/load_rookie_scale.py            # dry run
    build/venv/bin/python3 build/load_rookie_scale.py --apply
    build/venv/bin/python3 build/load_rookie_scale.py --sheet /path/to.xlsx

Needs openpyxl, which lives in build/venv (the system python3 does not have it).
"""

import argparse
import json
import os
import sys
import urllib.request
from pathlib import Path

DATA_DIR = Path(os.environ.get("NBS_DATA_DIR", "/var/lib/nothing-but-stats"))
OUT_FILE = DATA_DIR / "rookie-scale.json"
BIOS_FILE = DATA_DIR / "player-bios.json"

SHEET_URL = os.environ.get(
    "POOPOO_SHEET_URL",
    "https://docs.google.com/spreadsheets/d/e/2PACX-1vTqyqqC0O1U9O-2uwmMk6UIhSO58ukTa5HpXaU_IOQa3SEW8bLK5Wpjh_KA4YWePDgT2BIdhPO6Mieu/pub?output=xlsx",
)

PICKS_PER_ROUND = 30   # § 7.1 ties this to the team count; the sheet is fixed at 30
YEARS_PER_DEAL = 5     # 4 contract years + the trailing RFA hold
FIRST_COL = 5          # column F, zero-indexed
PICK_NUM_COL = 11      # column L, present on newer tabs only


def season_key(draft_year: int, offset: int) -> str:
    """2026, 0 -> '26-27'. The league year a draft's Year 1 lands in is the one
    starting that calendar year."""
    start = (draft_year + offset) % 100
    return f"{start:02d}-{(start + 1) % 100:02d}"


def read_tab(wb, draft_year: int) -> list[list[int]]:
    name = f"{draft_year} Rookie Contracts"
    if name not in wb.sheetnames:
        raise KeyError(name)
    ws = wb[name]
    rows = list(ws.iter_rows(min_row=2, max_row=1 + PICKS_PER_ROUND,
                             max_col=PICK_NUM_COL + 1, values_only=True))
    if len(rows) != PICKS_PER_ROUND:
        raise ValueError(f"{name}: expected {PICKS_PER_ROUND} rows, got {len(rows)}")

    out = []
    for i, row in enumerate(rows, start=1):
        vals = row[FIRST_COL:FIRST_COL + YEARS_PER_DEAL]
        if any(v in (None, "") for v in vals):
            raise ValueError(f"{name}: pick {i} has a blank year: {vals}")
        # Row order is pick order. Where the tab also carries an explicit pick
        # column, make it prove that rather than trusting the ordering.
        stated = row[PICK_NUM_COL] if len(row) > PICK_NUM_COL else None
        if stated is not None and int(stated) != i:
            raise ValueError(f"{name}: row {i} is labelled pick {int(stated)}")
        out.append([int(round(float(v))) for v in vals])

    firsts = [r[0] for r in out]
    if firsts != sorted(firsts, reverse=True):
        raise ValueError(f"{name}: Year 1 figures are not descending by pick")
    return out


def ratio_report(picks: list[list[int]]) -> list[str]:
    """Check the Y5 hold against Y4 for § 3.10's direction.

    The carve-out is 250% or 300% of the final rookie year, and which one you
    get turns on the same EAPS comparison as every other Full Bird hold: the
    *higher* multiplier belongs to the *lower* salary (190% at-or-below EAPS,
    150% above). So as pick number rises and salary falls, the multiplier may
    only rise. A table where the top picks take 300% and the late ones 250% has
    the comparison backwards, which is easy to do by hand and invisible in the
    dollar figures themselves. Returns [] when the table is consistent.
    """
    problems, ratios = [], []
    for i, row in enumerate(picks, start=1):
        y4, y5 = row[3], row[4]
        r = round(y5 / y4, 3) if y4 else 0
        ratios.append(r)
        if r not in (2.5, 3.0):
            problems.append(f"pick {i}: Y5 is {r}x Y4 — expected 2.5x or 3.0x")
    if problems:
        return problems
    for i in range(1, len(ratios)):
        if ratios[i] < ratios[i - 1]:
            return [f"pick {i} takes {ratios[i - 1]}x but pick {i + 1} takes {ratios[i]}x — "
                    f"the multiplier falls as salary falls, which is § 3.10 backwards",
                    f"(this table splits at pick {i}/{i + 1}; 2025 and 2026 split at 2/3)"]
    return []


def verify_against_bios(draft_year: int, picks: list[list[int]]) -> tuple[int, list[str]]:
    """Cross-check the table against contracts already on the books.

    This is the guard the sheet's own header row can't provide: if the seasons
    were shifted or the tab re-keyed, a signed pick's stored salaries stop
    lining up and this reports it instead of writing a wrong table.
    """
    if not BIOS_FILE.exists():
        return 0, ["player-bios.json not found — skipped verification"]
    bios = json.loads(BIOS_FILE.read_text())
    checked, problems = 0, []
    for slug, bio in bios.items():
        if bio.get("draft_year") != draft_year or bio.get("draft_round") != 1:
            continue
        num = bio.get("draft_pick")
        salaries = bio.get("salaries") or {}
        if not num or not (1 <= num <= PICKS_PER_ROUND) or bio.get("type") == "draft-rights":
            continue
        expected = picks[num - 1]
        for offset, amount in enumerate(expected):
            key = season_key(draft_year, offset)
            stored = salaries.get(key)
            if stored is None:
                continue  # signed to something other than the scale; not this script's call
            stored_n = int(str(stored).replace("$", "").replace(",", ""))
            if abs(stored_n - amount) > 1:
                problems.append(f"{slug} (pick {num}) {key}: sheet ${amount:,} vs bio ${stored_n:,}")
        checked += 1
    return checked, problems


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true", help="write the file (default is a dry run)")
    ap.add_argument("--sheet", help="path to an already-downloaded xlsx")
    ap.add_argument("--years", help="comma-separated draft years (default: every tab found)")
    args = ap.parse_args()

    try:
        import openpyxl
    except ImportError:
        print("openpyxl missing — run with build/venv/bin/python3", file=sys.stderr)
        return 2

    path = args.sheet
    if not path:
        path = "/tmp/nbn-rookie-sheet.xlsx"
        print(f"Downloading sheet -> {path}")
        urllib.request.urlretrieve(SHEET_URL, path)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if args.years:
        years = [int(y) for y in args.years.split(",")]
    else:
        years = sorted(int(n.split()[0]) for n in wb.sheetnames
                       if n.endswith("Rookie Contracts") and n.split()[0].isdigit())
    if not years:
        print("No '{year} Rookie Contracts' tabs found", file=sys.stderr)
        return 1

    scale = json.loads(OUT_FILE.read_text()) if OUT_FILE.exists() else {}
    failures = []
    for year in years:
        try:
            picks = read_tab(wb, year)
        except (KeyError, ValueError) as e:
            print(f"  {year}: SKIPPED — {e}")
            failures.append(year)
            continue
        checked, problems = verify_against_bios(year, picks)
        status = f"{len(picks)} picks, verified against {checked} signed contract(s)"
        if problems:
            print(f"  {year}: MISMATCH — {status}")
            for p in problems[:10]:
                print(f"      {p}")
            failures.append(year)
            continue
        ratios = ratio_report(picks)
        if ratios:
            print(f"  {year}: MISMATCH — § 3.10 hold multiplier looks inverted")
            for r in ratios:
                print(f"      {r}")
            failures.append(year)
            continue
        print(f"  {year}: OK — {status}")
        print(f"      seasons {season_key(year, 0)} .. {season_key(year, YEARS_PER_DEAL - 1)}"
              f" · pick 1 Y1 ${picks[0][0]:,} · pick 30 Y1 ${picks[-1][0]:,}")
        scale[str(year)] = picks

    # Each draft year is an independent table, so one bad tab must not hold the
    # others hostage — a year that verified is written, a year that didn't is
    # left untouched and named. The non-zero exit is what makes the skip loud.
    if not args.apply:
        print(f"\nDry run. Re-run with --apply to write {OUT_FILE}")
    else:
        OUT_FILE.write_text(json.dumps(scale, indent=2) + "\n")
        print(f"\nWrote {OUT_FILE} ({len(scale)} draft years)")

    if failures:
        print(f"\n{len(failures)} year(s) failed verification and were NOT loaded: "
              f"{failures}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
